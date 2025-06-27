from fastapi import FastAPI, HTTPException
from fastapi.responses import JSONResponse
from pydantic import BaseModel
from openpyxl import load_workbook
from datetime import datetime
import tempfile, base64, io

app = FastAPI()

# Constantes
tiempo_por_rol = {
    "Especialista SEO": {"Plata": 5, "Gold": 10, "Platinum": 12},
    "Analista SEO": {"Plata": 4, "Gold": 4, "Platinum": 6},
    "Técnico SEO": {"Plata": 4, "Gold": 0, "Platinum": 4},
    "Redactor SEO": {"Plata": 8, "Gold": 12, "Platinum": 15},
    "Community Manager": {"Plata": 12, "Gold": 20, "Platinum": 28},
    "Diseñador Senior": {"Plata": 6, "Gold": 10, "Platinum": 0},
    "Diseñador Junior": {"Plata": 4, "Gold": 6, "Platinum": 10},
    "Videógrafo": {"Plata": 6, "Gold": 6, "Platinum": 8},
    "Editor de Video": {"Plata": 4, "Gold": 6, "Platinum": 10},
    "Copywriter": {"Plata": 4, "Gold": 6, "Platinum": 8},
    "Director Creativo": {"Plata": 0, "Gold": 0, "Platinum": 4}
}
HORAS_MENSUALES = 160

# Modelo para recibir JSON
class AsignacionRequest(BaseModel):
    plan: str
    archivo_base64: str

@app.post("/asignar_plan")
def asignar_plan(req: AsignacionRequest):
    plan = req.plan
    if plan not in ["Plata", "Gold", "Platinum"]:
        raise HTTPException(status_code=400, detail="Tipo de plan inválido.")

    # Decodificar archivo base64
    try:
        contenido_binario = base64.b64decode(req.archivo_base64)
        archivo_excel = io.BytesIO(contenido_binario)
        wb = load_workbook(archivo_excel)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error al procesar Excel: {str(e)}")

    ws_ocu = wb["Ocupación"]
    ws_asig = wb["Asignaciones"] if "Asignaciones" in wb.sheetnames else wb.create_sheet("Asignaciones")
    ws_disp = wb["Disponibilidad"] if "Disponibilidad" in wb.sheetnames else wb.create_sheet("Disponibilidad")

    ocupacion = {}
    mes_actual = datetime.now().month
    for row in ws_ocu.iter_rows(min_row=2, values_only=True):
        rol, horas, mes = row
        if mes != mes_actual:
            horas = 0
        ocupacion[rol] = horas

    asignacion_binaria = {}
    for rol in tiempo_por_rol:
        req_horas = tiempo_por_rol[rol][plan]
        total = ocupacion.get(rol, 0) + req_horas
        if total <= HORAS_MENSUALES:
            asignacion_binaria[rol] = 1 if req_horas > 0 else 0
            ocupacion[rol] = total
        else:
            asignacion_binaria[rol] = 0

    for rol in tiempo_por_rol:
        if tiempo_por_rol[rol][plan] > 0 and asignacion_binaria[rol] == 0:
            raise HTTPException(status_code=409, detail=f"No hay disponibilidad suficiente para {rol}")

    next_row = ws_asig.max_row + 1 if ws_asig.max_row > 1 else 2
    if next_row == 2:
        ws_asig.append(["Proyecto", "Plan"] + list(tiempo_por_rol.keys()))
    fila = [next_row - 1, plan] + [asignacion_binaria[rol] for rol in tiempo_por_rol.keys()]
    ws_asig.append(fila)

    for i, rol in enumerate(tiempo_por_rol.keys(), start=2):
        ws_ocu[f"B{i}"] = ocupacion[rol]
        ws_ocu[f"C{i}"] = mes_actual

    disponibilidad = {"Plata": 999, "Gold": 999, "Platinum": 999}
    for plan_tipo in ["Plata", "Gold", "Platinum"]:
        min_posibles = 999
        for rol in tiempo_por_rol:
            req = tiempo_por_rol[rol][plan_tipo]
            if req > 0:
                disponibles = max(0, (HORAS_MENSUALES - ocupacion[rol]) // req)
                min_posibles = min(min_posibles, disponibles)
        disponibilidad[plan_tipo] = min_posibles

    ws_disp.delete_rows(1, ws_disp.max_row)
    ws_disp.append(["Plan", "Proyectos Posibles Este Mes"])
    for p in ["Platinum", "Gold", "Plata"]:
        ws_disp.append([p, disponibilidad[p]])

    # Guardar a binario y codificar de nuevo
    output_stream = io.BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)
    base64_resultado = base64.b64encode(output_stream.read()).decode("utf-8")

    return {
        "mensaje": f"Proyecto {next_row - 1} asignado correctamente.",
        "ocupacion_actual": ocupacion,
        "disponibilidad": disponibilidad,
        "archivo_actualizado_base64": base64_resultado
    }
